import csv
import io
import logging
import uuid
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime

from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone

from accounts.models import Profile
from billing.models import (
    ConsumptionRecord,
    RechargeRecord,
    UtilityBill,
    Wallet,
)
from billing.services.ledger_service import LedgerService
from data_center.models import DataTask, ImportRowError
from data_center.services.task_service import TaskService
from dormitory.models import RoomAssignment

logger = logging.getLogger(__name__)


def _money(value) -> Decimal:
    return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


class BaseImportHandler:
    name = 'base'
    label = '基础'
    required_fields = []
    optional_fields = []
    idempotency_fields = []

    @classmethod
    def all_fields(cls):
        return cls.required_fields + cls.optional_fields

    @classmethod
    def field_labels(cls):
        return {f: f for f in cls.all_fields()}

    @classmethod
    def validate_row(cls, row: dict, row_num: int) -> list:
        errors = []
        for f in cls.required_fields:
            val = row.get(f)
            if val is None or str(val).strip() == '':
                errors.append(f'必填字段「{f}」为空')
        return errors

    @classmethod
    def get_idempotency_key(cls, row: dict) -> str:
        parts = [str(row.get(f, '')).strip() for f in cls.idempotency_fields]
        return '|'.join(parts)

    @classmethod
    def check_idempotent(cls, idem_key: str) -> bool:
        return False

    @classmethod
    def import_row(cls, row: dict, operator: str) -> object:
        raise NotImplementedError


class UserImportHandler(BaseImportHandler):
    name = 'users'
    label = '用户'
    required_fields = ['username', 'student_id']
    optional_fields = ['email', 'phone', 'role', 'password']
    idempotency_fields = ['username', 'student_id']

    @classmethod
    def field_labels(cls):
        return {
            'username': '用户名',
            'student_id': '学号',
            'email': '邮箱',
            'phone': '手机号',
            'role': '角色(student/admin)',
            'password': '初始密码',
        }

    @classmethod
    def validate_row(cls, row: dict, row_num: int) -> list:
        errors = super().validate_row(row, row_num)
        role = str(row.get('role', '')).strip()
        if role and role not in ('student', 'admin'):
            errors.append('角色必须是 student 或 admin')
        phone = str(row.get('phone', '')).strip()
        if phone and len(phone) < 6:
            errors.append('手机号格式不正确')
        return errors

    @classmethod
    def check_idempotent(cls, idem_key: str) -> bool:
        parts = idem_key.split('|')
        if len(parts) < 2:
            return False
        username, student_id = parts[0], parts[1]
        if not username:
            return False
        if User.objects.filter(username=username).exists():
            return True
        if student_id and Profile.objects.filter(student_id=student_id).exists():
            return True
        return False

    @classmethod
    @transaction.atomic
    def import_row(cls, row: dict, operator: str) -> User:
        username = str(row['username']).strip()
        student_id = str(row.get('student_id', '')).strip() or None
        email = str(row.get('email', '')).strip()
        phone = str(row.get('phone', '')).strip() or None
        role = str(row.get('role', '')).strip() or 'student'
        raw_pwd = str(row.get('password', '')).strip() or uuid.uuid4().hex[:8]

        user, created = User.objects.get_or_create(
            username=username,
            defaults={
                'email': email,
                'password': make_password(raw_pwd),
            },
        )
        if not created:
            if email and not user.email:
                user.email = email
                user.save(update_fields=['email'])

        profile, _ = Profile.objects.get_or_create(user=user)
        if student_id and not profile.student_id:
            profile.student_id = student_id
        if phone and not profile.phone:
            profile.phone = phone
        if role:
            profile.role = role
        profile.save()

        Wallet.objects.get_or_create(user=user)
        return user


class RechargeImportHandler(BaseImportHandler):
    name = 'recharges'
    label = '充值记录'
    required_fields = ['username', 'amount', 'channel']
    optional_fields = ['student_id', 'operator', 'remark', 'order_no']
    idempotency_fields = ['username', 'amount', 'order_no']

    @classmethod
    def field_labels(cls):
        return {
            'username': '用户名',
            'student_id': '学号（用于定位用户）',
            'amount': '充值金额',
            'channel': '渠道(alipay/wechat/bank)',
            'operator': '操作人',
            'remark': '备注',
            'order_no': '业务单号（幂等键）',
        }

    @classmethod
    def validate_row(cls, row: dict, row_num: int) -> list:
        errors = super().validate_row(row, row_num)
        try:
            amount = Decimal(str(row.get('amount', '0')))
            if amount <= 0:
                errors.append('金额必须大于 0')
        except (InvalidOperation, ValueError):
            errors.append('金额格式不正确')
        channel = str(row.get('channel', '')).strip()
        if channel and channel not in ('alipay', 'wechat', 'bank'):
            errors.append('渠道必须是 alipay/wechat/bank 之一')
        return errors

    @classmethod
    def check_idempotent(cls, idem_key: str) -> bool:
        parts = idem_key.split('|')
        if len(parts) < 3:
            return False
        order_no = parts[2].strip()
        if order_no:
            if RechargeRecord.objects.filter(remark__icontains=order_no).exists():
                return True
        return False

    @classmethod
    @transaction.atomic
    def import_row(cls, row: dict, operator: str) -> RechargeRecord:
        username = str(row.get('username', '')).strip()
        student_id = str(row.get('student_id', '')).strip()
        amount = _money(row['amount'])
        channel = str(row['channel']).strip()
        row_operator = str(row.get('operator', '')).strip() or operator
        remark = str(row.get('remark', '')).strip()
        order_no = str(row.get('order_no', '')).strip()

        user = None
        if username:
            user = User.objects.filter(username=username).first()
        if not user and student_id:
            user = User.objects.filter(profile__student_id=student_id).first()
        if not user:
            raise ValueError(f'找不到用户: username={username}, student_id={student_id}')

        if order_no:
            existing = RechargeRecord.objects.filter(
                user=user,
                amount=amount,
                remark__icontains=order_no,
            ).first()
            if existing:
                return existing

        full_remark = remark + (f' [order_no:{order_no}]' if order_no else '')

        return LedgerService.create_recharge(
            user=user,
            amount=amount,
            channel=channel,
            operator=row_operator,
            remark=full_remark,
        )


class ConsumptionImportHandler(BaseImportHandler):
    name = 'consumptions'
    label = '消费记录'
    required_fields = ['username', 'category', 'usage', 'unit_price']
    optional_fields = ['student_id', 'meter_value', 'operator', 'remark']
    idempotency_fields = ['username', 'category', 'usage', 'unit_price', 'remark']

    @classmethod
    def field_labels(cls):
        return {
            'username': '用户名',
            'student_id': '学号（用于定位用户）',
            'category': '类型(water/electricity)',
            'usage': '用量',
            'unit_price': '单价',
            'meter_value': '当前表底',
            'operator': '操作人',
            'remark': '备注',
        }

    @classmethod
    def validate_row(cls, row: dict, row_num: int) -> list:
        errors = super().validate_row(row, row_num)
        try:
            usage = Decimal(str(row.get('usage', '0')))
            if usage <= 0:
                errors.append('用量必须大于 0')
        except (InvalidOperation, ValueError):
            errors.append('用量格式不正确')
        try:
            price = Decimal(str(row.get('unit_price', '0')))
            if price <= 0:
                errors.append('单价必须大于 0')
        except (InvalidOperation, ValueError):
            errors.append('单价格式不正确')
        category = str(row.get('category', '')).strip()
        if category and category not in ('water', 'electricity'):
            errors.append('类型必须是 water 或 electricity')
        return errors

    @classmethod
    def check_idempotent(cls, idem_key: str) -> bool:
        parts = idem_key.split('|')
        if len(parts) < 5:
            return False
        return False

    @classmethod
    @transaction.atomic
    def import_row(cls, row: dict, operator: str) -> ConsumptionRecord:
        username = str(row.get('username', '')).strip()
        student_id = str(row.get('student_id', '')).strip()
        category = str(row['category']).strip()
        usage = _money(row['usage'])
        unit_price = _money(row['unit_price'])
        meter_value = row.get('meter_value')
        meter_decimal = _money(meter_value) if meter_value not in (None, '', '0') else None
        row_operator = str(row.get('operator', '')).strip() or operator
        remark = str(row.get('remark', '')).strip()

        user = None
        if username:
            user = User.objects.filter(username=username).first()
        if not user and student_id:
            user = User.objects.filter(profile__student_id=student_id).first()
        if not user:
            raise ValueError(f'找不到用户: username={username}, student_id={student_id}')

        return LedgerService.create_consumption(
            user=user,
            category=category,
            usage=usage,
            unit_price=unit_price,
            meter_value=meter_decimal,
            operator=row_operator,
            remark=remark,
        )


class UtilityBillImportHandler(BaseImportHandler):
    name = 'utility_bills'
    label = '水电账单'
    required_fields = ['username', 'category', 'period_start', 'period_end', 'total_amount', 'due_date']
    optional_fields = [
        'student_id', 'usage', 'unit_price', 'previous_reading', 'current_reading',
        'status', 'operator', 'remark',
    ]
    idempotency_fields = ['username', 'category', 'period_start', 'period_end']

    @classmethod
    def field_labels(cls):
        return {
            'username': '用户名',
            'student_id': '学号',
            'category': '类型(water/electricity)',
            'period_start': '周期开始(YYYY-MM-DD)',
            'period_end': '周期结束(YYYY-MM-DD)',
            'usage': '用量',
            'unit_price': '单价',
            'previous_reading': '上期表底',
            'current_reading': '本期表底',
            'total_amount': '应缴金额',
            'due_date': '缴费截止日期(YYYY-MM-DD)',
            'status': '状态(pending/paid/void)',
            'operator': '操作人',
            'remark': '备注',
        }

    @classmethod
    def validate_row(cls, row: dict, row_num: int) -> list:
        errors = super().validate_row(row, row_num)
        for df in ('period_start', 'period_end', 'due_date'):
            v = str(row.get(df, '')).strip()
            if v:
                try:
                    datetime.strptime(v, '%Y-%m-%d')
                except ValueError:
                    errors.append(f'{df} 格式错误，应为 YYYY-MM-DD')
        try:
            amt = Decimal(str(row.get('total_amount', '0')))
            if amt < 0:
                errors.append('应缴金额不能为负')
        except (InvalidOperation, ValueError):
            errors.append('应缴金额格式不正确')
        category = str(row.get('category', '')).strip()
        if category and category not in ('water', 'electricity'):
            errors.append('类型必须是 water 或 electricity')
        status = str(row.get('status', '')).strip()
        if status and status not in ('pending', 'paid', 'void'):
            errors.append('状态必须是 pending/paid/void')
        return errors

    @classmethod
    def check_idempotent(cls, idem_key: str) -> bool:
        parts = idem_key.split('|')
        if len(parts) < 4:
            return False
        username, category, ps, pe = parts[0], parts[1], parts[2], parts[3]
        user = User.objects.filter(username=username).first()
        if not user:
            return False
        try:
            ps_d = datetime.strptime(ps, '%Y-%m-%d').date()
            pe_d = datetime.strptime(pe, '%Y-%m-%d').date()
        except ValueError:
            return False
        return UtilityBill.objects.filter(
            user=user, category=category, period_start=ps_d, period_end=pe_d,
        ).exists()

    @classmethod
    @transaction.atomic
    def import_row(cls, row: dict, operator: str) -> UtilityBill:
        username = str(row.get('username', '')).strip()
        student_id = str(row.get('student_id', '')).strip()
        category = str(row['category']).strip()
        period_start = datetime.strptime(str(row['period_start']).strip(), '%Y-%m-%d').date()
        period_end = datetime.strptime(str(row['period_end']).strip(), '%Y-%m-%d').date()
        total_amount = _money(row['total_amount'])
        due_date = datetime.strptime(str(row['due_date']).strip(), '%Y-%m-%d').date()
        usage = _money(row.get('usage', 0)) if row.get('usage') not in (None, '', '0') else Decimal('0')
        unit_price = _money(row.get('unit_price', 0)) if row.get('unit_price') not in (None, '', '0') else Decimal('0')
        previous_reading = _money(row.get('previous_reading', 0)) if row.get('previous_reading') not in (None, '', '0') else Decimal('0')
        current_reading = _money(row.get('current_reading', 0)) if row.get('current_reading') not in (None, '', '0') else Decimal('0')
        status = str(row.get('status', '')).strip() or UtilityBill.STATUS_PENDING
        row_operator = str(row.get('operator', '')).strip() or operator
        remark = str(row.get('remark', '')).strip()

        user = None
        if username:
            user = User.objects.filter(username=username).first()
        if not user and student_id:
            user = User.objects.filter(profile__student_id=student_id).first()
        if not user:
            raise ValueError(f'找不到用户: username={username}, student_id={student_id}')

        existing = UtilityBill.objects.filter(
            user=user, category=category, period_start=period_start, period_end=period_end,
        ).first()
        if existing:
            return existing

        assignment = RoomAssignment.objects.filter(
            user=user, unbound_at__isnull=True,
        ).select_related('room').first()
        if not assignment:
            raise ValueError(f'用户未绑定房间，无法创建水电账单。请先在宿舍管理中绑定房间。')

        bill_no = f'UB{timezone.now().strftime("%Y%m%d%H%M%S")}{uuid.uuid4().hex[:6].upper()}'

        return UtilityBill.objects.create(
            bill_no=bill_no,
            user=user,
            room=assignment.room,
            category=category,
            period_start=period_start,
            period_end=period_end,
            usage=usage,
            unit_price=unit_price,
            previous_reading=previous_reading,
            current_reading=current_reading,
            base_amount=total_amount,
            total_amount=total_amount,
            due_date=due_date,
            status=status,
            operator=row_operator,
            remark=remark,
        )


IMPORT_HANDLERS = {
    DataTask.DATA_TYPE_USERS: UserImportHandler,
    DataTask.DATA_TYPE_RECHARGES: RechargeImportHandler,
    DataTask.DATA_TYPE_CONSUMPTIONS: ConsumptionImportHandler,
    DataTask.DATA_TYPE_UTILITY_BILLS: UtilityBillImportHandler,
}


class ImportService:
    @staticmethod
    def get_handler(data_type: str):
        handler = IMPORT_HANDLERS.get(data_type)
        if not handler:
            raise ValueError(f'不支持的数据类型: {data_type}')
        return handler

    @staticmethod
    def parse_file(file_path: str, file_format: str) -> list:
        rows = []
        if file_format == 'csv':
            with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cleaned = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
                    rows.append(cleaned)
        else:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                header = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        header = [str(c).strip() if c is not None else f'col{i}' for i, c in enumerate(row)]
                        continue
                    if all(c in (None, '') for c in row):
                        continue
                    data = {}
                    for j, val in enumerate(row):
                        if j < len(header):
                            data[header[j]] = '' if val is None else str(val)
                    rows.append(data)
            except ImportError:
                raise ValueError('系统未安装 openpyxl，无法解析 XLSX 文件，请使用 CSV 格式或联系管理员。')
        return rows

    @staticmethod
    def preview(task: DataTask, file_path: str) -> dict:
        handler = ImportService.get_handler(task.data_type)
        rows = ImportService.parse_file(file_path, task.file_format)

        result = {
            'total_rows': len(rows),
            'headers': handler.all_fields(),
            'field_labels': handler.field_labels(),
            'rows': [],
            'error_count': 0,
            'warning_count': 0,
            'sample_data': [],
        }

        for idx, row in enumerate(rows, start=2):
            errors = handler.validate_row(row, idx)
            idem_key = handler.get_idempotency_key(row)
            is_dup = handler.check_idempotent(idem_key) if idem_key else False

            row_result = {
                'row_number': idx,
                'data': row,
                'errors': errors,
                'warnings': [],
                'is_idempotent_skip': is_dup,
            }
            if is_dup:
                row_result['warnings'].append('该行数据已存在，正式导入时将跳过（幂等控制）')
                result['warning_count'] += 1

            if errors:
                result['error_count'] += 1
            result['rows'].append(row_result)

            if idx <= 11:
                result['sample_data'].append(row_result)

        return result

    @staticmethod
    def execute_import(task: DataTask, file_path: str, corrected_rows: dict | None = None) -> dict:
        handler = ImportService.get_handler(task.data_type)
        corrected_rows = corrected_rows or {}
        rows = ImportService.parse_file(file_path, task.file_format)

        TaskService.mark_running(task)
        task.total_rows = len(rows)
        task.save(update_fields=['total_rows'])

        success_count = 0
        failed_count = 0
        skipped_count = 0
        error_rows = []

        ImportRowError.objects.filter(task=task).delete()

        for idx, raw_row in enumerate(rows, start=2):
            try:
                row = corrected_rows.get(str(idx), raw_row) if corrected_rows else raw_row
                errors = handler.validate_row(row, idx)

                idem_key = handler.get_idempotency_key(row)
                is_dup = handler.check_idempotent(idem_key) if idem_key else False

                if is_dup:
                    skipped_count += 1
                    ImportRowError.objects.create(
                        task=task,
                        row_number=idx,
                        row_data=row,
                        error_messages=['该行数据已存在，已跳过（幂等控制）'],
                        is_idempotent_skip=True,
                    )
                elif errors:
                    failed_count += 1
                    ImportRowError.objects.create(
                        task=task,
                        row_number=idx,
                        row_data=row,
                        error_messages=errors,
                    )
                    error_rows.append({
                        'row_number': idx,
                        **row,
                        '错误信息': '; '.join(errors),
                    })
                else:
                    try:
                        handler.import_row(row, task.operator.username)
                        success_count += 1
                    except Exception as e:
                        failed_count += 1
                        err_msg = str(e)
                        ImportRowError.objects.create(
                            task=task,
                            row_number=idx,
                            row_data=row,
                            error_messages=[err_msg],
                        )
                        error_rows.append({
                            'row_number': idx,
                            **row,
                            '错误信息': err_msg,
                        })

                progress = int((idx - 1) / len(rows) * 90) + 5
                if idx % 5 == 0 or idx == len(rows) + 1:
                    TaskService.update_progress(
                        task, progress,
                        success_rows=success_count,
                        failed_rows=failed_count,
                        skipped_rows=skipped_count,
                    )
            except Exception as e:
                logger.exception('导入行处理异常 row=%s', idx)
                failed_count += 1
                ImportRowError.objects.create(
                    task=task,
                    row_number=idx,
                    row_data=raw_row,
                    error_messages=[f'系统错误: {e}'],
                )

        if error_rows:
            headers = handler.all_fields() + ['错误信息']
            error_path = TaskService.save_csv_result(
                f'{task.data_type}_import_{task.id}',
                error_rows,
                headers,
                is_error=True,
            )
            task.error_file_path = error_path

        if failed_count == 0:
            TaskService.mark_success(
                task,
                success_rows=success_count,
                failed_rows=failed_count,
                skipped_rows=skipped_count,
            )
        elif success_count == 0:
            TaskService.mark_failed(task, f'全部 {failed_count} 行导入失败，请下载错误明细查看原因。')
            task.failed_rows = failed_count
            task.skipped_rows = skipped_count
            task.save(update_fields=['failed_rows', 'skipped_rows', 'error_file_path', 'updated_at'])
        else:
            TaskService.mark_partial(
                task,
                success_rows=success_count,
                failed_rows=failed_count,
                skipped_rows=skipped_count,
            )

        return {
            'total': len(rows),
            'success': success_count,
            'failed': failed_count,
            'skipped': skipped_count,
        }
